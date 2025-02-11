from rest_framework import status
from .models import ChildrenTempTable, CoreServiceChildren, CoreServiceFamily, CoreServiceChildrenMedicalContact, CoreServiceChildrenAllergies, CoreServiceClassrooms, CoreServiceChildrenEnrollment, StaffTempTable, Staff, App_user, Branch_users
from enjoey_api.models import BranchTable
from .serializers import ChildrenTempTableSerializer, CoreServiceChildrenTableSerializer, CoreServiceChildrenMedicalContactTableSerializer, CoreServiceChildrenAllergiesTableSerializer, CoreServiceChildrenEnrollmentTableSerializer, CoreServiceFamilyTableSerializer, CoreServiceClassroomsTableSerializer, StaffTempTableSerializer
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.decorators import api_view
from rest_framework import status
import csv
import datetime
from datetime import datetime
from django.db.models import Max
from openpyxl import load_workbook
from django.db import transaction
from .functions import get_staff_profile_image, generate_random_cognito_password, create_cognito_user

class UploadChildrenCSVData(APIView):
    def get(self, request, *args, **kwargs):
        try:
            records = ChildrenTempTable.objects.all()
            serializer = ChildrenTempTableSerializer(records, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')

        try:
            raw_data = file.read().decode('utf-8-sig')
            raw_lines = raw_data.splitlines()
            data_lines = raw_lines[4:]

            csv_data = csv.DictReader(data_lines)
            def convert_bool(value):
                if value == 'Yes':
                    return True
                elif value == 'No':
                    return False
                else:
                    return False
            
            def format_date(date_str):
                formats_to_try = ["%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d"]
                
                for date_format in formats_to_try:
                    try:
                        date_obj = datetime.strptime(date_str, date_format)
                        return date_obj.strftime("%d-%m-%Y")
                    except ValueError:
                        continue
                
                # Handle invalid date formats here
                return None
            
            csv_to_model_mapping = {
                "No": "no",
                "Name": "name",
                "Foreign Name": "foreignName",
                "Student ID (BC Number)": "studentID",
                "Gender": "gender",
                "Dob": "dob",
                "Age": "age",
                "Race": "race",
                "Citizenship": "citizenship",
                "Residential Status": "residentialStatus",
                "Religion": "religion",
                "Mother Tongue language": "motherTonguelanguage",
                "Child of a staff": "childOfStaff",
                "Household Income": "householdIncome",
                "Subsidy Type": "subsidyType",
                "Other Residential Remarks": "otherResidentialRemarks",
                "Siblings 1": "siblings1",
                "Siblings Relationship 1": "siblingsRelationship1",
                "Siblings 2": "siblings2",
                "Siblings Relationship 2": "siblingsRelationship2",
                "Siblings 3": "siblings3",
                "Siblings Relationship 3": "siblingsRelationship3",
                "Siblings 4": "siblings4",
                "Siblings Relationship 4": "siblingsRelationship4",
                "Medical Conditions": "medicalConditions",
                "Vaccination History": "vaccinationHistory",
                "Allergy History (Food/Medication)": "allergyHistory",
                "Special Diet": "specialDiet",
                "Special Needs": "specialNeeds",
                "Family Doctor Number": "familyDoctorNum",
                "Admission Date": "admissionDate",
                "Withdrawal date": "withdrawalDate",
                "Program Type": "programType",
                "Class name": "className",
                "Class Session": "classSession",
                "Class level": "classLevel",
                "Transportation Number": "transportationNum",
                "Payment Method": "paymentMethod1",
                "Bank Name": "bankName1",
                "Bank Account Holder": "bankAccountHolder1",
                "Bank Account Code": "bankAccountCode1",
                "Branch Code": "branchCode1",
                "Bank Account Number": "bankAccountNumber1",
                "Payment Method Approval Date": "paymentMethodApprovalDate1",
                "Attention To": "attentionTo1",
                "Payee ID": "payeeID",
                "Amount": "amount",
                "Payment Method": "paymentMethod2",
                "Bank Name": "bankName2",
                "Bank Account Holder": "bankAccountHolder2",
                "Bank Account Code": "bankAccountCode2",
                "Branch Code": "branchCode2",
                "Bank Account Number": "bankAccountNumber2",
                "Secondary Payee ID": "secondaryPayeeID",
                "Payment Method Approval Date": "paymentMethodApprovalDate2",
                "Attention To": "attentionTo2",
                "Payment Method": "paymentMethod3",
                "Bank Name": "bankName3",
                "Bank Account Holder": "bankAccountHolder3",
                "Bank Account Code": "bankAccountCode3",
                "Branch Code": "branchCode3",
                "Bank Account No.": "bankAccountNo",
                "Payment Method Approval Reference": "paymentMethodApprovalReference",
                "Payment Method Approval Date": "paymentMethodApprovalDate3",
                "Deposit Option": "depositOption",
                "Parent Relationship 1": "parentRelationship1",
                "Name 1": "name1",
                "Email 1": "email1",
                "Mobile No. 1": "mobileNo1",
                "Phone No. 1": "phoneNo1",
                "NRIC /ID 1": "NRIC1",
                "Race 1": "race1",
                "Citizenship 1": "citizenship1",
                "Occupation 1": "occupation1",
                "Main contact 1": "mainContact1",
                "Authorised Pick Up Person 1": "authorisedPickUpPerson1",
                "Email Invoice Receipt 1": "emailInvoiceReceipt1",
                "Email checkin 1": "emailCheckin1",
                "Parent Relationship 2": "parentRelationship2",
                "Name 2": "name2",
                "Email 2": "email2",
                "Mobile No. 2": "mobileNo2",
                "Phone No. 2": "phoneNo2",
                "NRIC /ID 2": "NRIC2",
                "Race 2": "race2",
                "Citizenship 2": "citizenship2",
                "Occupation 2": "occupation2",
                "Main contact 2": "mainContact2",
                "Authorised Pick Up Person 2": "authorisedPickUpPerson2",
                "Email Invoice Receipt 2": "emailInvoiceReceipt2",
                "Email checkin 2": "emailCheckin2",
                "Parent Relationship 3": "parentRelationship3",
                "Name 3": "name3",
                "Email 3": "email3",
                "Mobile No. 3": "mobileNo3",
                "Phone No. 3": "phoneNo3",
                "NRIC /ID 3": "NRIC3",
                "Race 3": "race3",
                "Citizenship 3": "citizenship3",
                "Occupation 3": "occupation3",
                "Main contact 3": "mainContact3",
                "Authorised Pick Up Person 3": "authorisedPickUpPerson3",
                "Email Invoice Receipt 3": "emailInvoiceReceipt3",
                "Email checkin 3": "emailCheckin3",
                "Block no. M": "blocknoM",
                "Building M": "buildingM",
                "Street M": "streetM",
                "Unit M": "unitM",
                "Postal Code M": "postalCodeM",
                "Transport Option M": "transportOptionM",
                "Block no. F": "blocknoF",
                "Building F": "buildingF",
                "Street F": "streetF",
                "Unit F": "unitF",
                "Postal Code F": "postalCodeF",
                "Transport Option F": "transportOptionF",
                "Address Line 1": "addressLine1",
                "Address Line 2": "addressLine2",
                "Postal Code RA1": "postalCodeRA1",
                "Transport Option RA1": "transportOptionRA1",
                "Relationship 1": "relationship1EC",
                "Name 1 EC": "name1EC",
                "Mobile Number 1": "mobileNumber1EC",
                "Phone Number 1": "phoneNumber1EC",
                "Email 1 EC": "email1EC",
            }

            max_badge_no = ChildrenTempTable.objects.aggregate(Max('badgeNo')).get('badgeNo__max')
            if max_badge_no:
                next_badge_no = int(max_badge_no) + 1
            else:
                next_badge_no = 10001  # Start from 10001 if no badge number exists

            # List of columns to remove
            columns_to_remove = [
                'School name', 'PT10 B6', 'Export date',
                'CHILD BASIC INFORMATION', 'Siblings', 'Health Information',
                'School Information', 'Fees Payment Primary', 'Fees Payment Secondary',
                'Refund Payment', 'Parent 1', 'Parent 2', 'Parent 3',
                'Residential Address (Mother)', 'Residential Address (Father)', 
                'Residential Address 1', 'Emergency contact 1'
            ]

            errors = []
            successful_uploads = 0
            failed_uploads = 0

            def insert_temp_table(row):
                row = {key: value for key, value in row.items() if key not in columns_to_remove}
                row = {csv_to_model_mapping.get(key, key): value for key, value in row.items()}
                row = {key.lstrip('\ufeff'): value for key, value in row.items()}
                row = {key: (value if value.strip() else 'null') for key, value in row.items()}
                row['Dob'] = format_date(row.get('Dob', ''))
                row['childOfStaff'] = convert_bool(row.get('childOfStaff', 'No'))
                row['mainContact1'] = convert_bool(row.get('mainContact1', 'No'))
                row['authorisedPickUpPerson1'] = convert_bool(row.get('authorisedPickUpPerson1', 'No'))
                row['emailInvoiceReceipt1'] = convert_bool(row.get('emailInvoiceReceipt1', 'No'))
                row['emailCheckin1'] = convert_bool(row.get('emailCheckin1', 'No'))
                row['mainContact2'] = convert_bool(row.get('mainContact1', 'No'))
                row['authorisedPickUpPerson2'] = convert_bool(row.get('authorisedPickUpPerson1', 'No'))
                row['emailInvoiceReceipt2'] = convert_bool(row.get('emailInvoiceReceipt1', 'No'))
                row['emailCheckin2'] = convert_bool(row.get('emailCheckin1', 'No'))
                row['mainContact3'] = convert_bool(row.get('mainContact1', 'No'))
                row['authorisedPickUpPerson3'] = convert_bool(row.get('authorisedPickUpPerson1', 'No'))
                row['emailInvoiceReceipt3'] = convert_bool(row.get('emailInvoiceReceipt1', 'No'))
                row['emailCheckin3'] = convert_bool(row.get('emailCheckin1', 'No'))
                row['badgeNo'] = next_badge_no

                serializer = ChildrenTempTableSerializer(data=row)
                if serializer.is_valid():
                    serializer.save()
                    return serializer.instance  # Return the model instance
                else:
                    errors.append(serializer.errors)
                    return None

            for row in csv_data:
                temp_record = insert_temp_table(row)
                if temp_record:
                    successful_uploads += 1
                else:
                    failed_uploads += 1
                # populate_core_service_children_table(temp_record)

            return Response({
                'message': 'CSV data imported successfully',
                'successful_uploads': successful_uploads,
                'failed_uploads': failed_uploads,
                'errors': errors
            }, status=status.HTTP_200_OK)
        
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class MigrateIntoCoreServiceChildrenCoreServiceFamilyTable(APIView):    
    def post(self, request, *args, **kwargs):
        badgeNo = request.data.get('badgeNo')
        className = request.data.get('className')
        filters = {}
        if badgeNo != "All": filters['badgeNo'] = badgeNo
        if className != "All": filters['className'] = className
        records = ChildrenTempTable.objects.filter(**filters)
        total_records = records.count()
        migrated_records = 0
        failed_records = 0
        duplicated_records = 0
        progress_updates = []
        errors = []

        if total_records == 0:
            return Response({"message": "No records to migrate"}, status=status.HTTP_200_OK)

        # Family Name Formatting
        def format_name(name):
            return ' '.join([part.capitalize() for part in name.split()])

        def get_first_name(full_name):
            if full_name == "null" or full_name is None or not full_name.strip():  # Skip records with "null", None or empty full_name
                return ''
            
            names = full_name.split()
            formatted_names = [format_name(name) for name in names]
            
            if len(formatted_names) == 1:
                return formatted_names[0]
            elif len(formatted_names) == 2:
                return formatted_names[0]
            elif len(formatted_names) == 3:
                return ' '.join(formatted_names[1:])
            elif len(formatted_names) == 4:
                return ' '.join(formatted_names[2:])
            return ''  # Return empty if no condition is met

        def get_last_name(full_name):
            if full_name == "null" or full_name is None or not full_name.strip():  # Skip records with "null", None or empty full_name
                return ''
            
            names = full_name.split()
            formatted_names = [format_name(name) for name in names]
            
            if len(formatted_names) == 1:
                return None
            elif len(formatted_names) == 2:
                return formatted_names[1]
            elif len(formatted_names) == 3:
                return formatted_names[0]
            elif len(formatted_names) == 4:
                return formatted_names[1]
            return ''  # Return empty if no condition is met

        # Family Phone Formatting
        def format_phone_number(phone_number, country_code='+60'):
            clean_number = ''.join(char for char in phone_number if char.isdigit())
            if not clean_number.startswith('0'):
                clean_number = '0' + clean_number
            clean_number = country_code + clean_number[1:]
            return f'{clean_number[:5]}-{clean_number[5:]}'
        
        def get_mobile_phone(mobile, phone):
            if mobile and mobile != "null":
                return format_phone_number(mobile)
            elif phone and phone != "null":
                return format_phone_number(phone)
            return "null"

        # Family Address Formatting
        def format_address(relationship, record):
            if relationship == "mother":
                address = getattr(record, 'streetM', '') or ''
                state = "Selangor"
                country = "Malaysia"
                postcode = getattr(record, 'postalCodeM', '') or ''
            elif relationship == "father":
                address = getattr(record, 'streetF', '') or ''
                state = "Selangor"
                country = "Malaysia"
                postcode = getattr(record, 'postalCodeF', '') or ''
            else:
                address_line1 = getattr(record, 'addressLine1', '') or ''
                address_line2 = getattr(record, 'addressLine2', '') or ''
                address = address_line1 + (' ' if address_line1 and address_line2 else '') + address_line2
                state = "Selangor"
                country = "Malaysia"
                postcode = getattr(record, 'postalCodeRA1', '') or ''
            return address, state, country, postcode

        try:
            for index, record in enumerate(records):
                if CoreServiceChildren.objects.filter(birthCertNo=record.studentID).exists():
                    duplicated_records += 1
                    progress_updates.append({
                        'migrated_records': migrated_records,
                        'duplicated_records': duplicated_records,
                        'failed_records': failed_records,
                        'progress': ((index + 1) / total_records) * 100
                    })
                    print("child", duplicated_records)
                    continue
                core_service_children_data = {
                    'fullName': record.name,
                    'birthCertNo': record.studentID,
                    'birthDate': record.dob,
                    'birthCountry': record.citizenship,
                    'ethnicity': record.race,
                    'religion': record.religion,
                    'gender': record.gender,
                    'age': record.age
                }
                serializer = CoreServiceChildrenTableSerializer(data=core_service_children_data)
                if serializer.is_valid():
                    serializer.save()
                    migrated_records += 1
                else:
                    failed_records += 1
                    errors.append(serializer.errors)

                for i in range(1, 4):
                    print("run")
                    name_key = f'name{i}'
                    mobile_key = f'mobileNo{i}'
                    phone_key = f'phoneNo{i}'
                    email_key = f'email{i}'
                    birthIC_key = f'NRIC{i}'
                    birthCountry_key = f'citizenship{i}'
                    occupation_key = f'occupation{i}'
                    relationship_key = f'parentRelationship{i}'
                    allow_pickup_key = f'authorisedPickUpPerson{i}'
                    billable_key = f'emailInvoiceReceipt{i}'
                    primary_key = f'mainContact{i}'

                    full_name = getattr(record, name_key, None)
                    birth_ic = getattr(record, birthIC_key, None)
                    email = getattr(record, email_key, None)
                    get_mobile = getattr(record, mobile_key, None)
                    get_phone = getattr(record, phone_key, None)
                    phone_no = get_mobile_phone(get_mobile, get_phone)
                    relationship = getattr(record, relationship_key, None)

                    if not full_name or full_name.strip() == "null":
                        failed_records += 1
                        continue

                    if birth_ic:
                        # Check for duplicate by birthIC if it's available
                        if CoreServiceFamily.objects.filter(birthIC=birth_ic).exists():
                            duplicated_records += 1
                            progress_updates.append({
                                'migrated_records': migrated_records,
                                'duplicated_records': duplicated_records,
                                'failed_records': failed_records,
                                'progress': ((index + 1) / total_records) * 100
                            })
                            continue  # Skip to the next family record if it's a duplicate
                    else:
                        # Check for duplicates by name and email if birthIC is null
                        if CoreServiceFamily.objects.filter(fullName=full_name, email=email).exists():
                            duplicated_records += 1
                            progress_updates.append({
                                'migrated_records': migrated_records,
                                'duplicated_records': duplicated_records,
                                'failed_records': failed_records,
                                'progress': ((index + 1) / total_records) * 100
                            })
                            continue  # Skip to the next family record if it's a duplicate

                    first_name = get_first_name(full_name)
                    last_name = get_last_name(full_name)
                    address, state, country, postcode = format_address(relationship, record)

                    core_service_family_data = {
                        'firstName': first_name,
                        'lastName': last_name,
                        'phone': phone_no,
                        'address': address,
                        'state': state,
                        'country': country,
                        'postcode': postcode,
                        'email': email,
                        'birthIC': getattr(record, birthIC_key, None),
                        'birthCountry': getattr(record, birthCountry_key, None),
                        'occupation': getattr(record, occupation_key, None),
                        'relationship': relationship,
                        'isAllowedPickup': getattr(record, allow_pickup_key, None),
                        'isBillable': getattr(record, billable_key, None),
                        'isPrimary': getattr(record, primary_key, None),
                    }
                    serializer = CoreServiceFamilyTableSerializer(data=core_service_family_data)
                    if serializer.is_valid():
                        serializer.save()
                        migrated_records += 1
                    else:
                        failed_records += 1
                        errors.append(serializer.errors)

                progress_updates.append({
                    'migrated_records': migrated_records,
                    'duplicated_records': duplicated_records,
                    'failed_records': failed_records,
                    'progress': ((index + 1) / total_records) * 100
                })
            return Response({
                'total_records': total_records,
                'migrated_records': migrated_records,
                'failed_records': failed_records,
                'duplicated_records': duplicated_records,
                'progress_updates': progress_updates,
                'errors': errors
            }, status=status.HTTP_200_OK)
        
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class MigrateIntoCoreServiceFamilyTable(APIView):
    def post(self, request, *args, **kwargs):
        badgeNo = request.data.get('badgeNo')
        className = request.data.get('className')
        filters = {}
        if badgeNo != "All": filters['badgeNo'] = badgeNo
        if className != "All": filters['className'] = className
        records = ChildrenTempTable.objects.filter(**filters)
        total_records = records.count()
        migrated_records = 0
        failed_records = 0
        duplicated_records = 0
        progress_updates = []
        errors = []

        if total_records == 0:
            return Response({"message": "No records to migrate"}, status=status.HTTP_200_OK)
        
        # Name Formatting
        def format_name(name):
            return ' '.join([part.capitalize() for part in name.split()])

        def get_first_name(full_name):
            if full_name == "null" or full_name is None or not full_name.strip():  # Skip records with "null", None or empty full_name
                return ''
            
            names = full_name.split()
            formatted_names = [format_name(name) for name in names]
            
            if len(formatted_names) == 1:
                return formatted_names[0]
            elif len(formatted_names) == 2:
                return formatted_names[0]
            elif len(formatted_names) == 3:
                return ' '.join(formatted_names[1:])
            elif len(formatted_names) == 4:
                return ' '.join(formatted_names[2:])
            return ''  # Return empty if no condition is met

        def get_last_name(full_name):
            if full_name == "null" or full_name is None or not full_name.strip():  # Skip records with "null", None or empty full_name
                return ''
            
            names = full_name.split()
            formatted_names = [format_name(name) for name in names]
            
            if len(formatted_names) == 1:
                return None
            elif len(formatted_names) == 2:
                return formatted_names[1]
            elif len(formatted_names) == 3:
                return formatted_names[0]
            elif len(formatted_names) == 4:
                return formatted_names[1]
            return ''  # Return empty if no condition is met

        # Phone Formatting
        def format_phone_number(phone_number, country_code='+60'):
            clean_number = ''.join(char for char in phone_number if char.isdigit())
            if not clean_number.startswith('0'):
                clean_number = '0' + clean_number
            clean_number = country_code + clean_number[1:]
            return f'{clean_number[:5]}-{clean_number[5:]}'
        
        def get_mobile_phone(mobile, phone):
            if mobile and mobile != "null":
                return format_phone_number(mobile)
            elif phone and phone != "null":
                return format_phone_number(phone)
            return "null"

        # Address Formatting
        def format_address(relationship, record):
            if relationship == "mother":
                address = getattr(record, 'streetM', '') or ''
                state = "Selangor"
                country = "Malaysia"
                postcode = getattr(record, 'postalCodeM', '') or ''
            elif relationship == "father":
                address = getattr(record, 'streetF', '') or ''
                state = "Selangor"
                country = "Malaysia"
                postcode = getattr(record, 'postalCodeF', '') or ''
            else:
                address_line1 = getattr(record, 'addressLine1', '') or ''
                address_line2 = getattr(record, 'addressLine2', '') or ''
                address = address_line1 + (' ' if address_line1 and address_line2 else '') + address_line2
                state = "Selangor"
                country = "Malaysia"
                postcode = getattr(record, 'postalCodeRA1', '') or ''
            return address, state, country, postcode

        try:
            for index, record in enumerate(records):
                for i in range(1, 4):
                    name_key = f'name{i}'
                    mobile_key = f'mobileNo{i}'
                    phone_key = f'phoneNo{i}'
                    email_key = f'email{i}'
                    birthIC_key = f'NRIC{i}'
                    birthCountry_key = f'citizenship{i}'
                    occupation_key = f'occupation{i}'
                    relationship_key = f'parentRelationship{i}'
                    allow_pickup_key = f'authorisedPickUpPerson{i}'
                    billable_key = f'emailInvoiceReceipt{i}'
                    primary_key = f'mainContact{i}'

                    full_name = getattr(record, name_key, None)
                    birth_ic = getattr(record, birthIC_key, None)
                    get_mobile = getattr(record, mobile_key, None)
                    get_phone = getattr(record, phone_key, None)
                    phone_no = get_mobile_phone(get_mobile, get_phone)
                    relationship = getattr(record, relationship_key, None)

                    if not full_name or full_name.strip() == "null":
                        failed_records += 1
                        continue

                    if CoreServiceFamily.objects.filter(birthIC=birth_ic).exists():
                        duplicated_records += 1
                        progress_updates.append({
                            'migrated_records': migrated_records,
                            'duplicated_records': duplicated_records,
                            'failed_records': failed_records,
                            'progress': ((index + 1) / total_records) * 100
                        })
                        continue

                    first_name = get_first_name(full_name)
                    last_name = get_last_name(full_name)
                    address, state, country, postcode = format_address(relationship, record)

                    core_service_family_data = {
                        'firstName': first_name,
                        'lastName': last_name,
                        'phone': phone_no,
                        'address': address,
                        'state': state,
                        'country': country,
                        'postcode': postcode,
                        'email': getattr(record, email_key, None),
                        'birthIC': getattr(record, birthIC_key, None),
                        'birthCountry': getattr(record, birthCountry_key, None),
                        'occupation': getattr(record, occupation_key, None),
                        'relationship': relationship,
                        'isAllowedPickup': getattr(record, allow_pickup_key, None),
                        'isBillable': getattr(record, billable_key, None),
                        'isPrimary': getattr(record, primary_key, None),
                    }
                    serializer = CoreServiceFamilyTableSerializer(data=core_service_family_data)
                    if serializer.is_valid():
                        serializer.save()
                        migrated_records += 1
                    else:
                        failed_records += 1
                        errors.append(serializer.errors)

                    progress_updates.append({
                        'migrated_records': migrated_records,
                        'duplicated_records': duplicated_records,
                        'failed_records': failed_records,
                        'progress': ((index + 1) / total_records) * 100
                    })
            return Response({
                'total_records': total_records,
                'migrated_records': migrated_records,
                'failed_records': failed_records,
                'duplicated_records': duplicated_records,
                'progress_updates': progress_updates,
                'errors': errors
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class UploadStaffCSVData(APIView):
    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')

        try:
            workbook = load_workbook(file)
            sheet = workbook.active
            def format_boolean(value):
                if isinstance(value, bool):  # If it's already a boolean
                    return value
                if isinstance(value, str):  # If it's a string
                    value = value.strip().upper()  # Normalize the input (remove spaces and uppercase)
                    if value == "TRUE":
                        return True
                    elif value == "FALSE":
                        return False
                # Handle invalid inputs
                return None
            
            def format_date(date_input):
                if isinstance(date_input, datetime):  # Check if input is already a datetime object
                    return date_input.strftime("%Y-%m-%d")
                elif isinstance(date_input, str):  # If it's a string, try to parse it
                    formats_to_try = ["%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d", "%d-%b-%y"]
                    for date_format in formats_to_try:
                        try:
                            date_obj = datetime.strptime(date_input, date_format)
                            return date_obj.strftime("%Y-%m-%d")
                        except ValueError:
                            continue
                # Handle invalid inputs
                return None
            
            headers = [
                "schoolName", "branchName", "firstName", "lastName", "email",
                "phone", "gender", "dob", "profileImage", "address", "state",
                "country", "birthCountry", "postcode", "role", "isWebAccess",
                "isMobileAccess", "doj", "isExternal", "staffNRIC", "academyYear",
                "academyMonth", "startDate", "classroomName", "isPrimary",
                "branches", "isFranchiseStaff"
            ]

            csv_to_model_mapping = {
                "schoolName": "schoolName",
                "branchName": "branchName",
                "firstName": "firstName",
                "lastName": "lastName",
                "email": "email",
                "phone": "phone",
                "gender": "gender",
                "dob": "dob",
                "profileImage": "profileImage",
                "address": "address",
                "state": "state",
                "country": "country",
                "birthCountry": "birthCountry",
                "postcode": "postcode",
                "role": "role",
                "isWebAccess": "isWebAccess",
                "isMobileAccess": "isMobileAccess",
                "doj": "doj",
                "isExternal": "isExternal",
                "staffNRIC": "staffNRIC",
                "academyYear": "academyYear",
                "academyMonth": "academyMonth",
                "startDate": "startDate",
                "classroomName": "classroomName",
                "isPrimary": "isPrimary",
                "Branches": "branches",
                "IsFranchiseStaff": "isFranchiseStaff",
            }

            max_badge_no = StaffTempTable.objects.aggregate(Max('badgeNo')).get('badgeNo__max')
            if max_badge_no:
                next_badge_no = int(max_badge_no) + 1
            else:
                next_badge_no = 10001  # Start from 10001 if no badge number exists

            errors = []
            successful_uploads = 0
            failed_uploads = 0

            def insert_temp_table(row):
                print('hi', row)
                row = {csv_to_model_mapping.get(headers[i], headers[i]): value for i, value in enumerate(row)}
                row = {key: (value if value and str(value).strip() else 'null') for key, value in row.items()}
                row['dob'] = format_date(row.get('dob', ''))
                row['doj'] = format_date(row.get('doj', ''))
                row['startDate'] = format_date(row.get('startDate', ''))
                row['isPrimary'] = format_boolean(row.get('isPrimary', ''))
                row['badgeNo'] = next_badge_no
                
                profile_image = row.get('profileImage')
                if not profile_image or not hasattr(profile_image, 'read'):
                    row['profileImage'] = None 

                serializer = StaffTempTableSerializer(data=row)
                if serializer.is_valid():
                    serializer.save()
                    return serializer.instance  # Return the model instance
                else:
                    errors.append(serializer.errors)
                    return None

            for row in sheet.iter_rows(min_row=2):  # Start from the first row
                row_values = [cell.value for cell in row]  # Extract values
                temp_record = insert_temp_table(row_values)
                if temp_record:
                    successful_uploads += 1
                else:
                    failed_uploads += 1

            return Response({
                'message': 'CSV data imported successfully',
                'successful_uploads': successful_uploads,
                'failed_uploads': failed_uploads,
                'errors': errors
            }, status=status.HTTP_200_OK)
        
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class CreateTenantStaff(APIView):
    def post(self, request):
    # authenticated_user: TokenUser = request.user
    # print('request.data', request.POST)
    # if authenticated_user:
        try:
            success_count = 0
            error_records = []
            with transaction.atomic():
                temp_staff_records = StaffTempTable.objects.all()
                for temp_record in temp_staff_records:
                    try: 
                        if Staff.objects.filter(email=temp_record.email, staffNRIC=temp_record.staffNRIC).exists():
                                error_records.append({
                                    "firstName": temp_record.firstName,
                                    "lastName": temp_record.lastName,
                                    "staffNRIC": temp_record.staffNRIC,
                                    "error": "Duplicate staff record found."
                                })
                                continue
                        
                        # branch = Branch.objects.filter(name=temp_record.branchName).first()
                        # if not branch:
                        #     raise ValueError(f"Branch with name '{temp_record.branchName}' does not exist.")
                        
                        # classrooms = ClassRooms.objects.filter(
                        #     className=temp_record.className,
                        #     tenantId=temp_record.schoolName,
                        #     branchId=branch.id,
                        #     academicYear=temp_record.academicYear
                        # )
                        # if not classrooms.exists():
                        #     raise ValueError(f"No classroom found for className: '{temp_record.className}', "
                        #                     f"tenantId: '{temp_record.schoolName}', academicYear: '{temp_record.academicYear}'.")
                        
                        staff_model = Staff()
                        staff_model.tenantId = temp_record.schoolName
                        # staff_model.branchId = branch.id
                        staff_model.branchId = temp_record.branchName
                        staff_model.email = temp_record.email
                        staff_model.firstName = temp_record.firstName
                        staff_model.lastName = temp_record.lastName
                        staff_model.phone = temp_record.phone
                        staff_model.staffNRIC = temp_record.staffNRIC
                        staff_model.birthCountry = temp_record.birthCountry
                        staff_model.role = temp_record.role
                        staff_model.dob = temp_record.dob
                        staff_model.doj = temp_record.doj
                        staff_model.isExternal = temp_record.isExternal
                        staff_model.gender = temp_record.gender
                        staff_model.address = temp_record.address
                        staff_model.state = temp_record.state
                        staff_model.country = temp_record.country
                        staff_model.postcode = temp_record.postcode
                        staff_model.isActive = True
                        staff_model.isWebAccess = temp_record.isWebAccess
                        staff_model.isMobileAccess = temp_record.isMobileAccess
                        staff_model.creator = request.user
                        staff_model.profileImage=temp_record.profileImage if temp_record.profileImage else None
                        staff_model.save()
                        
                        app_user_model = App_user()
                        app_user_model.tenantId=staff_model.tenantId
                        app_user_model.email=staff_model.email
                        app_user_model.phone=staff_model.phone
                        app_user_model.firstName=staff_model.firstName
                        app_user_model.lastName=staff_model.lastName
                        app_user_model.roles=[staff_model.role]
                        app_user_model.isActive=True
                        app_user_model.created_by=request.user.id
                        app_user_model.save()

                        # Link App_user to Staff record
                        staff_model.userId = app_user_model.id
                        staff_model.isWebAccess = True
                        staff_model.save()
                        
                        # temp_password = generate_random_cognito_password()
                        # profile_image_url = get_staff_profile_image(staff_model)
                        # cognito_response = create_cognito_user(authenticated_user, app_user_model, temp_password, profile_image_url)

                        # if cognito_response.get('User'):
                        #     app_user_model.tempPassword = temp_password
                        #     app_user_model.save()
                        # else:
                        #     raise Exception(f"Failed to create Cognito user for staff: {staff.email}")
                        
                        # for classroom in classrooms:
                        #     classroom_staff = Classroom_staff()
                        #     classroom_staff.staff=staff_model
                        #     classroom_staff.classroom=classroom
                        #     classroom_staff.academicYear=temp_record.academicYear
                        #     classroom_staff.isPrimary=temp_record.isPrimary
                        #     classroom_staff.isActive=True
                        #     classroom_staff.save()
                        
                        # if staff_model.role in ['admin', 'management']:
                        #     branches = BranchTable.objects.filter(tenant_id=staff_model.tenantId)
                        #     for branch in branches:
                        #         branch_user = Branch_users()
                        #         branch_user.tenantId = staff_model.tenantId
                        #         branch_user.branch = branch
                        #         branch_user.user = app_user_model
                        #         branch_user.creator = request.user.id
                        #         branch_user.save()
                        # else:
                        #     branch = BranchTable.objects.filter(name=temp_record.branchName, tenant_id=staff_model.tenantId).first()
                        #     if not branch:
                        #         raise ValueError(f"Branch with name '{temp_record.branchName}' does not exist.")
                        #     branch_user = Branch_users()
                        #     branch_user.tenantId = staff_model.tenantId
                        #     branch_user.branch = branch
                        #     branch_user.user = app_user_model
                        #     branch_user.creator = request.user.id
                        #     branch_user.save()
                        
                        success_count += 1

                    except Exception as inner_error:
                        error_records.append({
                            "firstName": temp_record.firstName,
                            "lastName": temp_record.lastName,
                            "staffNRIC": temp_record.staffNRIC,
                            "error": str(inner_error)
                        })
                    
                temp_staff_records.delete()
                
            return Response({'message': 'Staff data successfully created.', 'success_count': success_count, 'errors': error_records}, status=status.HTTP_200_OK)

        except Exception as e:
            print(e)
            return Response({'message': str(e)}, status=status.HTTP_404_NOT_FOUND)
